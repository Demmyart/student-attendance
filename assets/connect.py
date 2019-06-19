#!/usr/bin/python
# -*- coding: utf-8
import sys  
import os
from flask import Flask, request, jsonify, Response, redirect, url_for
from flask_mysqldb import MySQL
# from requests import main
# sys.path.append("/diplom/")
from api import predict
import json
import datetime
from openpyxl import load_workbook
from werkzeug.utils import secure_filename


app = Flask(__name__)
# app.register_blueprint(main)

UPLOAD_FOLDER = 'photos/'
ALLOWED_EXTENSIONS = set(['jpg', 'png','jpeg', 'xls', 'xlsx'])

app.config['MYSQL_HOST'] = '127.0.0.1'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'university'
app.config['MYSQL_USE_UNICODE'] = 'True'
app.config['JSON_AS_ASCII'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
mysql = MySQL(app)



"""АВТОРИЗАЦИЯ"""
@app.route("/auth", methods=['GET', 'POST'])
def login():
    cursor = mysql.connection.cursor()
    data = request.get_json(force=True)
    username=data.get('username')
    password=data.get('password')
    cursor.execute("SELECT *,student.id as 'id',student.name as 'name',groups.name as 'namename' FROM student JOIN  \
    groups ON student.group_id = groups.id  \
    WHERE login = '{}' \
    UNION ALL SELECT *,teacher.id,teacher.name,department.name FROM teacher \
    JOIN department ON teacher.id_department = department.id \
    WHERE login = '{}';".format(username,username))
    query = cursor.fetchall()
    if query:
        if password == query[0][5]:
            row_headers=[x[0] for x in cursor.description]
            json_data=[]
            for result in query:
                json_data.append(dict(zip(row_headers,result)))
            response = {
                        'data': json_data,
                        'stat': 'success',
                        'message': 'Successfully logged in.',
                }
            return jsonify(response)
        else:
            response = {
                    'stat': 'Fail',
                    'message': 'Wrong password',
                    }
            return jsonify(response)
        return jsonify(response)
    else:
        response = {
                    'stat': 'Fail',
                    'message': 'Doesnt exist.',
                    }
        return jsonify(response)


"""РАСПИСАНИЕ ДЛЯ УЧИТЕЛЯ НА ДЕНЬ"""
@app.route("/schedule", methods=['GET', 'POST'])
def schedule():
    cursor = mysql.connection.cursor()
    get_date = request.get_json(force = True)
    id_teacher = get_date['id_teacher']
    day_name = datetime.datetime.strptime(get_date['date'], '%Y-%m-%d').strftime('%A')
    week_number = datetime.datetime.strptime(get_date['date'], '%Y-%m-%d').strftime('%V')
    # day_name = datetime.datetime.now().strftime('%A')
    # week_number = datetime.datetime.now().strftime('%V')
    if int(week_number) % 2 == 0:
        day_name1 = day_name+"-count"
    else: 
        day_name1 = day_name+"-uncount"
    cursor.execute("SELECT schedule.id,schedule.start,schedule.end,\
    schedule.day,teacher.name,schedule.id_teacher,teacher.surname,teacher.last_name,\
    lesson.name AS 'lesson_name',classroom.name as 'classroom_name', \
    GROUP_CONCAT(schedule_group.id_group) AS 'id_group' FROM schedule\
    JOIN teacher ON schedule.id_teacher=teacher.id\
    JOIN lesson ON schedule.id_lesson=lesson.id JOIN classroom ON \
    schedule.id_classroom=classroom.id\
    JOIN schedule_group ON schedule.id= schedule_group.id_schedule\
    WHERE id_teacher='{}' AND (schedule.day = '{}' \
    OR schedule.day ='{}') GROUP BY schedule.id;".format(id_teacher,day_name1,day_name))
    query= cursor.fetchall()
    row_headers=[x[0] for x in cursor.description]
    json_data=[]
    for result in query:
        json_data.append(dict(zip(row_headers,result)))
    return json.dumps(json_data, default=str,ensure_ascii = False)



"""СПИСОК СТУДЕНТОВ ПО ВЫБРАННОЙ ПАРЕ"""
@app.route("/lessoninfo", methods=['GET', 'POST'])
def lessoninfo():
    cursor = mysql.connection.cursor()
    get_info = request.get_json(force = True)
    id_groups = get_info['id_group'].split(',')
    id_schedule = get_info['id_schedule']
    json_data1=[]
    json_data=[]
    json_data1.append(dict({'id': id_schedule}))
    for i in id_groups:
        cursor.execute("SELECT student.name, student.surname,student.last_name, student.group_id,student.id as 'id' \
        FROM student WHERE student.group_id = '{}'".format(i))
        query= cursor.fetchall()
        row_headers=[x[0] for x in cursor.description]
        for result in query:
            json_data.append(dict(zip(row_headers,result)))
    for el in json_data:
        el['here'] = 0
    return jsonify(json_data,json_data1)



def allowed_filename(filename):
    return '.' in filename and filename.rsplit('.',1)[1] in ALLOWED_EXTENSIONS


"""ЗАГРУЗКА ФОТОГРАФИИ - В БАЗУ"""
@app.route("/sendphoto", methods=['GET', 'POST'])
def sendphoto():
    if request.method == 'POST':
        submitted_file = request.files['File']
        date = datetime.datetime.today().strftime('%Y-%m-%d')
        id_schedule = request.form['id_schedule']
        if submitted_file and allowed_filename(submitted_file.filename):
            filename = secure_filename(submitted_file.filename)
            submitted_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            cursor = mysql.connection.cursor()
            cursor.execute("INSERT INTO current_schedule \
            (id_schedule, date, photo) VALUES \
            ('{}','{}','{}');".format(id_schedule,date,submitted_file.filename))
            mysql.connection.commit()
            res = predict(submitted_file)
    return jsonify(res)

def str_to_bool(s):
    if s == 'true':
         return True
    elif s == 'false':
         return False

"""ОТМЕТКА О ПРИСУТСТВИИ В РУЧНУЮ"""
@app.route("/studentishere", methods=['GET', 'POST'])
def studentishere():
    cursor = mysql.connection.cursor()
    get_info = request.get_json(force = True)
    id_schedule = get_info['id_schedule']
    date = datetime.datetime.today().strftime('%Y-%m-%d')
    id_student = get_info['id']
    lessondate = get_info['date']
    flag = get_info['flag']
    # listnothere = get_info['listnothere'].split(',')
    json_data=[]

    # cursor.execute("INSERT INTO current_schedule (id_schedule, date) \
    # SELECT * FROM (SELECT '{}','{}') AS tmp \
    # WHERE NOT EXISTS ( \
    # SELECT date  FROM current_schedule WHERE date = '{}' \
    # ) LIMIT 1 SELECT ".format(id_schedule,date,lessondate))

    cursor.execute("SELECT id FROM current_schedule WHERE date = '{}' and id_schedule='{}'".format(lessondate,id_schedule))
    idforme = cursor.fetchone()
    if idforme:
        idforme = idforme[0]
        flagNum = 0 if flag == str_to_bool('false') else 1
        cursor.execute("INSERT INTO statistic (id_student, id_current_schedule,here) \
                VALUES ('{}','{}','{}')".format(id_student,idforme, flagNum))

        
        # if flag == str_to_bool('true'):
        #     cursor.execute("INSERT INTO statistic (id_student, id_current_schedule,here) \
        #             VALUES ('{}','{}',1)".format(id_student,idforme))
        # if flag == str_to_bool('false'):
        #     cursor.execute("INSERT INTO statistic (id_student, id_current_schedule,here) \
        #             VALUES ('{}','{}',0)".format(id_student,idforme))
        cursor.execute("SELECT id_student as 'id', here FROM statistic WHERE id_current_schedule = '{}' \
        and id_student = '{}'".format(idforme,id_student))
        query= cursor.fetchall()
        row_headers=[x[0] for x in cursor.description]
        for result in query:
            json_data.append(dict(zip(row_headers,result)))
    else:
        cursor.execute("INSERT INTO current_schedule (id_schedule, date) VALUES ('{}','{}')".format(id_schedule,lessondate))
        if flag == str_to_bool('true'):
            cursor.execute("INSERT INTO statistic (id_student, id_current_schedule,here) \
                    VALUES ('{}',LAST_INSERT_ID(),1)".format(id_student))
        if flag == str_to_bool('false'):
            cursor.execute("INSERT INTO statistic (id_student, id_current_schedule,here) \
                    VALUES ('{}',LAST_INSERT_ID(),0)".format(id_student))
        cursor.execute("SELECT id_student as 'id', here FROM statistic WHERE id_current_schedule = LAST_INSERT_ID()")
        query= cursor.fetchall()
        row_headers=[x[0] for x in cursor.description]
        for result in query:
            json_data.append(dict(zip(row_headers,result)))
    mysql.connection.commit()

        
    #     cursor.execute("IF EXISTS (SELECT * FROM statistic WHERE \
    #     id_student = '{}' AND id_current_schedule = LAST_INSERT_ID())  \
    #         BEGIN \
    #              UPDATE statistic SET here = 1 \
    #              WHERE id_student = '{}' AND id_current_schedule = LAST_INSERT_ID() \
    #         END \
    #     ELSE \
    #          BEGIN \
    #             INSERT INTO statistic (id_student, id_current_schedule,here) \
    #                 VALUES ('{}',LAST_INSERT_ID(),1)\
    #         END".format(id_student,id_student,id_student))

    # for i in listhere:
    #     cursor.execute("INSERT INTO statistic (id_student, \
    #     id_current_schedule,here) VALUES ('{}',LAST_INSERT_ID(),1)".format(i))
    # for i in listnothere:
    #     cursor.execute("INSERT INTO statistic (id_student, \
    #     id_current_schedule,here) VALUES ('{}',LAST_INSERT_ID(),0)".format(i))
    # for el in json_data:
    #     el['processing'] = str_to_bool('false')
    return jsonify(json_data)



"""ПОИСК - СПИСОК ПАР ПО ГРУППЕ"""
@app.route("/searchbygroups", methods=['GET', 'POST'])
def searchbygroups():
    cursor = mysql.connection.cursor()
    get_info = request.get_json(force = True)
    id_group = get_info['id_group']
    id_teacher = get_info['id_teacher']
    json_data=[]
    cursor.execute("SELECT id_schedule as 'id',groups.name,lesson.name as 'lesson_name',groups.id as 'id_group' \
    FROM schedule_group \
    JOIN groups ON schedule_group.id_group = groups.id \
    JOIN schedule ON schedule_group.id_schedule = schedule.id \
    JOIN lesson ON schedule.id_lesson = lesson.id\
    WHERE groups.name = '{}' && schedule.id_teacher = '{}' "
    .format( id_group,id_teacher))
    query= cursor.fetchall()
    row_headers=[x[0] for x in cursor.description]
    for result in query:
        json_data.append(dict(zip(row_headers,result)))
    return jsonify(json_data)


"""СТАТИСТИКА ГРУППЫ ПО ПРЕДМЕТУ"""
@app.route("/getstatbygroup", methods=['GET', 'POST'])
def getstatbygroup():
    cursor = mysql.connection.cursor()
    get_info = request.get_json(force = True)
    id_group = get_info['id_group']
    id_schedule = get_info['id_schedule']
    json_data=[]
    cursor.execute("SELECT statistic.id_student as 'id',student.name,student.surname,student.last_name,\
    statistic.here, \
    GROUP_CONCAT(statistic.here) AS 'stats',GROUP_CONCAT(statistic.id_current_schedule) AS 'schedules' FROM statistic JOIN student ON \
    student.id = statistic.id_student \
    JOIN current_schedule ON statistic.id_current_schedule = current_schedule.id \
    JOIN schedule ON schedule.id = current_schedule.id_schedule \
    WHERE schedule.id = '{}' AND student.group_id='{}' GROUP BY statistic.id_student"
    .format(id_schedule, id_group))
    query= cursor.fetchall()
    row_headers=[x[0] for x in cursor.description]
    for result in query:
        json_data.append(dict(zip(row_headers,result)))
    # json_datamy = json.dumps(json_data,ensure_ascii = False)
    # json_datamy1 = json.loads(json_datamy)
    for el in json_data:
        results = [int(i) for i in el['stats'].split(',')]
        percent = sum(results)/float(len(results))*100
        # json_data[i].append(dict({'total': percent}))
        el['total'] = percent
    return jsonify(json_data)


"""РАСПИСАНИЕ ДЛЯ CТУДЕНТА НА ДЕНЬ"""
@app.route("/schedulestudent", methods=['GET', 'POST'])
def schedulestudent():
    cursor = mysql.connection.cursor()
    get_date = request.get_json(force = True)
    id_group = get_date['id_group']
    day_name = datetime.datetime.strptime(get_date['date'], '%Y-%m-%d').strftime('%A')
    week_number = datetime.datetime.strptime(get_date['date'], '%Y-%m-%d').strftime('%V')
    # day_name = datetime.datetime.now().strftime('%A')
    # week_number = datetime.datetime.now().strftime('%V')
    if int(week_number) % 2 == 0:
        day_name1 = day_name+"-count"
    else: 
        day_name1 = day_name+"-uncount"
    cursor.execute("SELECT schedule.id as 'id',schedule.start,schedule.end,\
    schedule.day,teacher.name,schedule.id_teacher,teacher.surname,teacher.last_name,\
    lesson.name AS 'lesson_name',classroom.name as 'classroom_name' \
    FROM schedule_group\
    JOIN schedule ON schedule.id= schedule_group.id_schedule \
    JOIN teacher ON schedule.id_teacher=teacher.id\
    JOIN lesson ON schedule.id_lesson=lesson.id JOIN classroom ON \
    schedule.id_classroom=classroom.id\
    WHERE id_group='{}' AND (schedule.day = '{}' \
    OR schedule.day ='{}');".format(id_group,day_name1,day_name))
    query= cursor.fetchall()
    row_headers=[x[0] for x in cursor.description]
    json_data=[]
    for result in query:
        json_data.append(dict(zip(row_headers,result)))
    return json.dumps(json_data, default=str,ensure_ascii = False)





"""СТАТИСТИКА СТУДЕНТА ПО ПРЕДМЕТУ"""
@app.route("/getstatforstudent", methods=['GET', 'POST'])
def getstatforstudent():
    cursor = mysql.connection.cursor()
    get_info = request.get_json(force = True)
    id_student = get_info['id_student']
    id_schedule = get_info['id_schedule']
    json_data=[]
    cursor.execute("SELECT statistic.id_student as 'id',student.name,student.surname,student.last_name,\
    statistic.here, \
    GROUP_CONCAT(statistic.here) AS 'stats',GROUP_CONCAT(statistic.id_current_schedule) AS 'schedules' FROM statistic JOIN student ON \
    student.id = statistic.id_student \
    JOIN current_schedule ON statistic.id_current_schedule = current_schedule.id \
    JOIN schedule ON schedule.id = current_schedule.id_schedule \
    WHERE schedule.id = '{}' AND student.id='{}' GROUP BY statistic.id_student"
    .format(id_schedule, id_student))
    query= cursor.fetchall()
    row_headers=[x[0] for x in cursor.description]
    for result in query:
        json_data.append(dict(zip(row_headers,result)))
    # json_datamy = json.dumps(json_data,ensure_ascii = False)
    # json_datamy1 = json.loads(json_datamy)
    for el in  json_data:
        results = [int(i) for i in el['stats'].split(',')]
        percent = sum(results)/float(len(results))*100
        # json_data[i].append(dict({'total': percent}))
        el['total'] = percent
    return jsonify(json_data)




"""ОБЩАЯ СТАТИСТИКА ДЛЯ ПАНЕЛИ СБОКУ"""
@app.route("/getTotalStat", methods=['GET', 'POST'])
def getTotalStat():
    cursor = mysql.connection.cursor()
    json_data=[]
    cursor.execute("SELECT AVG(here) FROM statistic")
    query= cursor.fetchall()
    for result in query:
        json_data.append(result)
    return json.dumps(json_data, default=str)


"""АУДИТОРИИ"""
@app.route("/getByAud", methods=['GET', 'POST'])
def getByAud():
    cursor = mysql.connection.cursor()
    json_data=[]
    day = datetime.datetime.today().strftime('%Y-%m-%d')
    cursor.execute("SELECT SUM(here) AS 'quantity',classroom.name,schedule.id,schedule.start,schedule.end, GROUP_CONCAT(statistic.here) \
    AS 'here_values'  FROM statistic  \
    JOIN current_schedule ON statistic.id_current_schedule =current_schedule.id JOIN schedule \
    ON schedule.id= current_schedule.id_schedule JOIN classroom ON schedule.id_classroom = classroom.id \
    WHERE current_schedule.date = '{}' AND (CONCAT(DATE_FORMAT(NOW(), '%H:%i'), ':00') BETWEEN schedule.start AND schedule.end) GROUP BY classroom.name ".format(day))
    query= cursor.fetchall()
    row_headers=[x[0] for x in cursor.description]
    for result in query:
        json_data.append(dict(zip(row_headers,result)))
    return json.dumps(json_data, default=str)


"""ЗАГРУЗКА РАСПИСАНИЯ"""
@app.route("/sendSchedule", methods=['GET', 'POST'])
def sendSchedule():
    if request.method == 'POST':
        submitted_file = request.files['File']
        if submitted_file and allowed_filename(submitted_file.filename):
            filename = secure_filename(submitted_file.filename) 
            submitted_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            ffg = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            mylist = []
            listforgroups=[]
            wb = load_workbook(ffg)
            ws = wb['table']
            for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
                newlist=[]
                for cell in row[:-1]:
                    newlist.append(str(cell.value))
                mylist.append(newlist)
            cursor = mysql.connection.cursor()
            cursor.executemany("INSERT INTO schedule \
            (id_lesson,id_teacher,id_classroom,start,end,day) VALUES \
            (%s,%s,%s,%s,%s,%s)",mylist)
            mysql.connection.commit()
            # for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
            #     newlist=[]
            #     for cell in row[-1:]:
            #         newlist.append(str(cell.value))
            #     listforgroups.append(newlist)
            # cursor = mysql.connection.cursor()
            # cursor.executemany("INSERT INTO schedule_group \
            # (id_schedule,id_group) VALUES \
            # (LAST_INSERT_ID(),%s)",listforgroups)
            # mysql.connection.commit()
    return jsonify("success")


    
# @app.route("/pred", methods=['GET', 'POST'])
# def pred():
#     try:
#         submitted_file = request.files['File']
#         res = predict(submitted_file)
#         return jsonify(res)
#     except:
#         return jsonify('xyi')


@app.after_request
def add_header(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

if __name__ == "__main__":
    app.run(debug=True)